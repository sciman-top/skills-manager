#!/usr/bin/env python3
"""
Gmail Extraction + Archive Pipeline

Extracts emails from Gmail, routes to the correct /mnt/ace/<repo>/ location
based on sender domain routing rules, commits to each repo, then optionally
deletes from Gmail inbox.

Usage:
  uv run scripts/email/gmail-archive-extract.py --account ace --query "from:sandsig.com" --max 500 --delete
  uv run scripts/email/gmail-archive-extract.py --account personal --query "in:inbox" --max 1000 --delete
  uv run scripts/email/gmail-archive-extract.py --account skestates --query "from:familydollar.com" --dry-run

Routing rules: scripts/email/email-routing.yaml
"""

import argparse, base64, csv, io, json, os, re, shutil, subprocess, sys, time
from collections import Counter
from datetime import datetime
from pathlib import Path

# ============================================================
# PATHS
# ============================================================
SCRIPTS_DIR = Path(__file__).resolve().parent
ROUTING_FILE = SCRIPTS_DIR / "email-routing.yaml"
OAUTH_ENV = os.path.expanduser("~/.gmail-mcp/oauth-env.json")
ACE_BASE = Path("/mnt/ace")

# Legal deny list — check multiple locations
DENY_PATH = None
for p in [
    Path("/mnt/ace/.ace-knowledge/.legal-deny-list.yaml"),
    Path("/mnt/local-analysis/workspace-hub/.legal-deny-list.yaml"),
]:
    if p.exists():
        DENY_PATH = p
        break
if DENY_PATH is None:
    DENY_PATH = Path("/dev/null")

# ============================================================
# CONFIG
# ============================================================
def load_routing():
    """Simple YAML parser — no external deps needed."""
    rules = {}
    if not ROUTING_FILE.exists():
        return rules
    with open(ROUTING_FILE) as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith('#') or line == 'rules:':
                continue
            if ':' not in line:
                continue
            # Split on first ": " or ":"
            parts = line.split(':', 1)
            domain = parts[0].strip().strip('"').strip("'")
            val = parts[1].strip().strip('"').strip("'")
            # Remove inline comments
            if '  #' in val:
                val = val.split('  #')[0].strip()
            rules[domain] = val
    return rules

def load_oauth():
    with open(OAUTH_ENV) as f:
        return json.load(f)

ACCOUNTS = {
    "ace":       {"email": "vamsee.achanta@aceengineer.com",  "token": "~/.gmail-ace/credentials.json"},
    "personal":  {"email": "achantav@gmail.com",               "token": "~/.gmail-personal/credentials.json"},
    "skestates": {"email": "skestatesinc@gmail.com",           "token": "~/.gmail-skestates/credentials.json"},
}

# ============================================================
# AUTH
# ============================================================
import urllib.request, urllib.parse

def refresh_token(acct):
    oauth = load_oauth()
    cred_path = os.path.expanduser(ACCOUNTS[acct]["token"])
    with open(cred_path) as f:
        saved = json.load(f)
    data = urllib.parse.urlencode({
        "client_id": oauth["client_id"], "client_secret": oauth["client_secret"],
        "refresh_token": saved["refresh_token"], "grant_type": "refresh_token",
    }).encode()
    req = urllib.request.Request("https://oauth2.googleapis.com/token", data=data, method="POST")
    with urllib.request.urlopen(req, timeout=30) as resp:
        tokens = json.loads(resp.read().decode())
    saved.update(tokens)
    with open(cred_path, "w") as f:
        json.dump(saved, f)
    return tokens["access_token"]

# ============================================================
# GMAIL API
# ============================================================
def gmail_search(token, query, max_results=500):
    import urllib.request, urllib.parse
    enc = urllib.parse.quote(query)
    req = urllib.request.Request(
        f"https://gmail.googleapis.com/gmail/v1/users/me/messages?q={enc}&maxResults={max_results}",
        headers={"Authorization": f"Bearer {token}"})
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            return json.loads(resp.read().decode())
    except Exception as e:
        return None

def gmail_full(token, msg_id):
    import urllib.request
    req = urllib.request.Request(
        f"https://gmail.googleapis.com/gmail/v1/users/me/messages/{msg_id}?format=full",
        headers={"Authorization": f"Bearer {token}"})
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            return json.loads(resp.read().decode())
    except:
        return None

def gmail_get_attachment(token, msg_id, att_id):
    import urllib.request
    req = urllib.request.Request(
        f"https://gmail.googleapis.com/gmail/v1/users/me/messages/{msg_id}/attachments/{att_id}",
        headers={"Authorization": f"Bearer {token}"})
    try:
        with urllib.request.urlopen(req, timeout=120) as resp:
            return json.loads(resp.read().decode())
    except:
        return None

def gmail_delete(token, msg_id):
    import urllib.request
    req = urllib.request.Request(
        f"https://gmail.googleapis.com/gmail/v1/users/me/messages/{msg_id}",
        headers={"Authorization": f"Bearer {token}"}, method="DELETE")
    try:
        urllib.request.urlopen(req, timeout=10)
        return True
    except:
        return False

# ============================================================
# EXTRACTION HELPERS
# ============================================================
def extract_text_body(payload):
    def walk(part):
        if part.get("body", {}).get("data"):
            try:
                decoded = base64.urlsafe_b64decode(part["body"]["data"]).decode("utf-8", errors="replace")
                if part.get("mimeType") == "text/plain":
                    return decoded
                elif part.get("mimeType") == "text/html":
                    text = re.sub(r'<[^>]+>', '\n', decoded)
                    text = re.sub(r'\s{2,}', ' ', text)
                    text = re.sub(r'\n{3,}', '\n\n', text)
                    return text.strip()
            except:
                pass
        for sub in part.get("parts", []):
            r = walk(sub)
            if r:
                return r
        return ""
    return walk(payload)

def find_attachments(parts):
    atts = []
    for part in parts:
        fname = part.get("filename", "")
        mime = part.get("mimeType", "")
        body = part.get("body", {})
        aid = body.get("attachmentId", "")
        if fname and aid:
            atts.append({"filename": fname, "mimeType": mime, "attachmentId": aid,
                         "size": body.get("size", 0)})
        for sub in part.get("parts", []):
            atts.extend(find_attachments([sub]))
    return atts

def clean_name(name):
    return re.sub(r'[^\w.\-_]', '_', name)[:80]

def parse_spreadsheet(xlsx_path, sheet_dir):
    csv_paths = []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        for sn in wb.sheetnames:
            ws = wb[sn]
            csv_path = sheet_dir / f"{clean_name(sn)}.csv"
            with open(csv_path, "w", newline="") as f:
                w = csv.writer(f)
                for row in ws.iter_rows(values_only=True):
                    w.writerow([v if v is not None else "" for v in row])
            csv_paths.append(str(csv_path))
    except ImportError:
        pass
    except Exception as e:
        print(f"    WARNING: spreadsheet parse error: {e}")
    return csv_paths

# ============================================================
# LEGAL SCAN
# ============================================================
def legal_scan(text, deny_path):
    if not DENY_PATH.exists() or str(DENY_PATH) == "/dev/null":
        return []
    try:
        import yaml
        with open(DENY_PATH) as f:
            deny = yaml.safe_load(f)
    except:
        return []
    
    matches = []
    text_lower = text.lower()
    for item in deny.get("client_references", []):
        pattern = item["pattern"]
        cs = item.get("case_sensitive", False)
        search = text if cs else text_lower
        search_pat = pattern if cs else pattern.lower()
        if search_pat in search:
            matches.append((pattern, item.get("description", "")))
    return matches

# ============================================================
# REPO HELPERS
# ============================================================
def git_commit(repo_path, msg):
    """Git add + commit in target repo."""
    for cmd_args in [["git", "add", "-A"], ["git", "commit", "-m", msg]]:
        r = subprocess.run(cmd_args, cwd=str(repo_path), capture_output=True, text=True)
        if r.returncode != 0 and cmd_args[1] == "commit":
            return f"⚠️ commit failed: {r.stderr[:200]}"
    return "✅ committed"

# ============================================================
# MAIN
# ============================================================
def main():
    import yaml
    
    parser = argparse.ArgumentParser(description="Extract Gmail messages to repos")
    parser.add_argument("--account", required=True, choices=list(ACCOUNTS.keys()))
    parser.add_argument("--query", default="in:inbox -in:trash", help="Gmail search query")
    parser.add_argument("--max", type=int, default=500, help="Max messages")
    parser.add_argument("--delete", action="store_true", help="Delete from Gmail after archive")
    parser.add_argument("--dry-run", action="store_true", help="Show plan without saving")
    parser.add_argument("--no-sheets", action="store_true", help="Skip spreadsheet parsing")
    parser.add_argument("--force", action="store_true", help="Proceed through legal violations")
    args = parser.parse_args()
    
    # Load config
    routing = load_routing()
    token = refresh_token(args.account)
    
    print(f"\n{'='*70}")
    print(f"  EXTRACTING: {args.account} ({ACCOUNTS[args.account]['email']})")
    print(f"  QUERY:      {args.query}")
    print(f"  MAX:        {args.max}")
    if args.delete:  print(f"  MODE:       EXTRACT + DELETE")
    else:            print(f"  MODE:       EXTRACT ONLY")
    print(f"{'='*70}\n")
    
    # Search
    result = gmail_search(token, args.query, args.max)
    msgs = result.get("messages", []) if result else []
    print(f"Found {len(msgs)} messages\n")
    
    stats = {
        "routed": Counter(),      # repo -> count
        "attachments": 0,
        "spreadsheets": 0,
        "csv_files": 0,
        "deleted": 0,
        "legal_blocked": 0,
        "review_flagged": 0,
        "errors": 0,
        "repo_commits": Counter(),  # repo -> commits
    }
    pending_commits = {}  # repo -> [files added]
    
    for i, msg_stub in enumerate(msgs):
        try:
            detail = gmail_full(token, msg_stub["id"])
            if not detail:
                continue
            
            hdrs = {h["name"]: h["value"] for h in detail.get("payload", {}).get("headers", [])}
            from_hdr = hdrs.get("From", "")
            match = re.search(r'<([^>]+)>', from_hdr)
            email_addr = match.group(1).lower() if match else from_hdr.lower().strip()
            domain = email_addr.split("@")[-1] if "@" in email_addr else "unknown"
            subject = hdrs.get("Subject", "(no subject)")
            
            # Look up routing
            action = routing.get(domain) or routing.get("default", "achantas-data/docs/email/other")
            
            if action == "DELETE":
                if args.delete and not args.dry_run:
                    gmail_delete(token, msg_stub["id"])
                stats["deleted"] += 1
                stats["routed"]["deleted"] += 1
                print(f"  [{i+1}] DELETE: {domain} — {subject[:60]}")
                continue
            
            if action == "REVIEW":
                stats["review_flagged"] += 1
                stats["routed"]["review"] += 1
                print(f"  [{i+1}] REVIEW: {domain} — {subject[:60]}")
                # Still extract for review
                target_repo = "achantas-data"
                target_path = Path("/mnt/ace/achantas-data/docs/email/review")
            else:
                # "repo/path/subpath"
                parts = action.split("/")
                target_repo = parts[0]
                target_path = ACE_BASE / target_repo / "/".join(parts[1:])
            
            # Create target directories
            target_path.mkdir(parents=True, exist_ok=True)
            att_dir = target_path.parent / "attachments" if target_path.name != "attachments" else target_path
            sheet_dir = target_path.parent / "spreadsheets" if target_path.name != "spreadsheets" else target_path
            att_dir.mkdir(exist_ok=True)
            sheet_dir.mkdir(exist_ok=True)
            
            # Extract body
            body = extract_text_body(detail.get("payload", {}))
            
            # Find attachments
            atts = find_attachments(detail.get("payload", {}).get("parts", []))
            att_info = []
            csv_paths = []
            
            # Download attachments
            if atts:
                att_subdir = att_dir / clean_name(subject[:60])
                att_subdir.mkdir(exist_ok=True)
                for att in atts:
                    try:
                        att_data = gmail_get_attachment(token, msg_stub["id"], att["attachmentId"])
                        raw = base64.urlsafe_b64decode(att_data["data"])
                        att_path = att_subdir / clean_name(att["filename"])
                        if not args.dry_run:
                            with open(att_path, "wb") as f:
                                f.write(raw)
                        att_info.append({"filename": att["filename"], "mimeType": att["mimeType"],
                                        "size": att["size"]})
                        stats["attachments"] += 1
                        
                        if att["mimeType"].startswith("application/vnd") and not args.no_sheets:
                            xlsx = sheet_dir / clean_name(att["filename"])
                            if not args.dry_run:
                                shutil.copy2(att_path, xlsx)
                                parsed = parse_spreadsheet(str(xlsx), sheet_dir)
                                csv_paths.extend(parsed)
                                stats["csv_files"] += len(parsed)
                    except Exception as e:
                        print(f"    ⚠️ attachment error: {att['filename'][:40]}")
            
            # Legal scan
            full_text = subject + " " + body
            violations = legal_scan(full_text, DENY_PATH)
            if violations and not args.force:
                print(f"  [{i+1}] BLOCKED: {subject[:60]} — {[v[0] for v in violations[:3]]}")
                stats["legal_blocked"] += 1
                continue
            
            # Build markdown
            lines = [f"# {subject}", ""]
            lines.append("| Field | Value |")
            lines.append("|-------|-------|")
            lines.append(f"| **From** | {hdrs.get('From', '')} |")
            lines.append(f"| **To** | {hdrs.get('To', '')} |")
            if hdrs.get("Cc"):
                lines.append(f"| **CC** | {hdrs['Cc']} |")
            lines.append(f"| **Date** | {hdrs.get('Date', '')} |")
            if atts:
                lines.append(f"| **Attachments** | {len(atts)} files |")
            if csv_paths:
                lines.append(f"| **Parsed CSVs** | {len(csv_paths)} sheets |")
            lines.append("")
            
            if att_info:
                lines.append("## Attachments")
                lines.append("")
                for a in att_info:
                    lines.append(f"- 📎 `{a['filename']}` ({a['size']:,} bytes, {a['mimeType']})")
                lines.append("")
            
            if csv_paths:
                lines.append("## Extracted Spreadsheets")
                lines.append("")
                for cp in csv_paths:
                    lines.append(f"- 📊 `{os.path.basename(cp)}`")
                lines.append("")
            
            lines.append("## Body")
            lines.append("")
            lines.append(body)
            lines.append("")
            lines.append("---")
            lines.append(f"*Archived: {datetime.now().isoformat()}*")
            
            # Save markdown
            date_prefix = hdrs.get("Date", "")[:10].replace(",", "")
            slug = clean_name(f"{date_prefix}_{subject[:60]}")
            md_path = target_path / f"{slug}.md"
            if not args.dry_run:
                with open(md_path, "w", encoding="utf-8") as f:
                    f.write("\n".join(lines))
            
            # Track for repo commit
            repo_key = str(ACE_BASE / target_repo)
            if repo_key not in pending_commits:
                pending_commits[repo_key] = {"repo": target_repo, "count": 0, "files": set()}
            pending_commits[repo_key]["count"] += 1
            pending_commits[repo_key]["files"].add(str(md_path))
            
            # Delete from Gmail
            if args.delete and not args.dry_run:
                gmail_delete(token, msg_stub["id"])
                stats["deleted"] += 1
            
            stats["routed"][target_repo] += 1
            
            att_tag = f" +{len(atts)}📎" if atts else ""
            csv_tag = f" → {len(csv_paths)}csv" if csv_paths else ""
            print(f"  [{i+1}/{len(msgs):3d}] {target_repo + '/':30s} {subject[:50]}{att_tag}{csv_tag}")
            
            # Rate limit
            time.sleep(0.15)
            
        except Exception as e:
            print(f"  [{i+1}/{len(msgs):3d}] ERROR: {e}")
            stats["errors"] += 1
    
    # Commit to each repo
    print(f"\n{'='*70}")
    print(f"  COMMITTING TO REPOS")
    print(f"{'='*70}")
    for repo_path, info in pending_commits.items():
        repo_dir = Path(repo_path)
        msg = f"extract: {args.account} email — {info['count']} messages from {args.query[:50]}"
        result = git_commit(repo_dir, msg)
        stats["repo_commits"][info["repo"]] = 1
        print(f"  {info['repo']}: {result} — {info['count']} files")
    
    # Summary
    print(f"\n{'='*70}")
    print(f"  SUMMARY")
    print(f"{'='*70}")
    print(f"  Messages scanned:  {len(msgs)}")
    for repo, count in sorted(stats["routed"].items(), key=lambda x: -x[1]):
        print(f"  → {repo:30s} {count}")
    if stats["review_flagged"]:
        print(f"  → {'REVIEW flagged':30s} {stats['review_flagged']}")
    if stats["legal_blocked"]:
        print(f"  → {'BLOCKED (legal)':30s} {stats['legal_blocked']}")
    print(f"  Attachments:       {stats['attachments']}")
    print(f"  CSV sheets parsed: {stats['csv_files']}")
    print(f"  Deleted:           {stats['deleted']}")
    print(f"  Errors:            {stats['errors']}")

if __name__ == "__main__":
    main()
