"""
Validate xlsx sidecar data against authoritative .owr binary.

Loads each .owr with OrcFxAPI and compares frequencies, RAO amplitudes,
and added-mass values against the corresponding xlsx produced by the
queue post-processor. Run on licensed-win-1 where OrcFxAPI is available.

Usage:
    python scripts/solver/validate_xlsx_against_owr.py
"""
import OrcFxAPI
import openpyxl
import numpy as np
from pathlib import Path
import sys

FIXTURES = Path(r"digitalmodel\tests\fixtures\solver")
CASES = [
    ("test01_unit_box", "test01_unit_box.owr", "test01_unit_box.xlsx"),
    ("ellipsoid", "ellipsoid.owr", "ellipsoid.xlsx"),
]


def validate_case(name, owr_file, xlsx_file):
    owr_path = FIXTURES / owr_file
    xlsx_path = FIXTURES / xlsx_file

    # Load from OrcFxAPI
    d = OrcFxAPI.Diffraction()
    d.LoadResults(str(owr_path.resolve()))
    freq_hz = np.array(d.frequencies)
    freq_rad = freq_hz * 2 * np.pi
    sort_idx = np.argsort(freq_rad)
    freq_rad_sorted = freq_rad[sort_idx]
    headings = np.array(d.headings)
    raos = np.array(d.displacementRAOs)  # (nheading, nfreq, 6)

    n_freq = len(freq_hz)
    n_head = len(headings)

    # Load from xlsx
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["RAOs"]
    headers = [c.value for c in ws[1]]
    data_rows = list(ws.iter_rows(min_row=2, values_only=True))
    xlsx_freqs = np.array([r[0] for r in data_rows])

    print(f"\n=== {name} ===")
    print(f"  OrcFxAPI: {n_freq} freqs, {n_head} headings")
    print(f"  xlsx:     {len(xlsx_freqs)} freqs")

    # Compare frequencies
    freq_diff = np.max(np.abs(freq_rad_sorted - xlsx_freqs))
    freq_ok = freq_diff < 1e-6
    print(f"  Freq max diff: {freq_diff:.2e} rad/s  {'PASS' if freq_ok else 'FAIL'}")

    # Compare RAO amplitudes for surge (DOF 0) at first heading
    owr_surge = np.abs(raos[0, sort_idx, 0])
    surge_col_name = "Surge_Mag_H" + str(headings[0])
    surge_col = headers.index(surge_col_name)
    xlsx_surge = np.array([abs(float(r[surge_col] or 0)) for r in data_rows])

    amp_diff = np.max(np.abs(owr_surge - xlsx_surge))
    rel_diff = np.max(np.abs(owr_surge - xlsx_surge) / (owr_surge + 1e-30))
    amp_ok = rel_diff < 0.01
    print(f"  Surge amp max abs diff: {amp_diff:.2e}")
    print(f"  Surge amp max rel diff: {rel_diff * 100:.4f}%  {'PASS' if amp_ok else 'FAIL'}")

    # Compare added mass diagonal (surge-surge)
    am = np.array(d.addedMass)[sort_idx]
    ws_am = wb["AddedMass"]
    am_headers = [c.value for c in ws_am[1]]
    am_rows = list(ws_am.iter_rows(min_row=2, values_only=True))
    ss_col = am_headers.index("Surge_Surge")
    owr_ss = am[:, 0, 0]
    xlsx_ss = np.array([float(r[ss_col] or 0) for r in am_rows])
    am_diff = np.max(np.abs(owr_ss - xlsx_ss))
    am_ok = am_diff < 1e-6
    print(f"  AddedMass(1,1) max abs diff: {am_diff:.2e}  {'PASS' if am_ok else 'FAIL'}")

    return freq_ok and amp_ok and am_ok


all_pass = True
for name, owr, xlsx in CASES:
    if not validate_case(name, owr, xlsx):
        all_pass = False

print(f"\n{'=' * 40}")
print(f"OVERALL: {'ALL PASS' if all_pass else 'SOME FAILURES'}")
sys.exit(0 if all_pass else 1)
