"""Tests for XLSX formula extraction size limit increase (#1619).

Validates:
  - Formula extraction from small XLSX files
  - Large file handling up to 50MB (mocked)
  - Memory-efficient streaming via openpyxl read_only mode
  - Error handling for corrupt XLSX files
  - Size limit is correctly set to 50MB
"""

import io
import os
import sys
import tempfile
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest

REPO_ROOT = str(Path(__file__).resolve().parents[2])
if REPO_ROOT not in sys.path:
    sys.path.insert(0, REPO_ROOT)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _create_small_xlsx(formulas: bool = True) -> str:
    """Create a small XLSX file with optional formulas, return path."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Calculations"
    ws["A1"] = "Input_A"
    ws["B1"] = "Input_B"
    ws["C1"] = "Result"
    ws["A2"] = 10.0
    ws["B2"] = 20.0
    if formulas:
        ws["C2"] = "=A2+B2"
        ws["C3"] = "=A2*B2"
        ws["C4"] = "=SUM(A2:B2)"
    else:
        ws["C2"] = 30.0

    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    wb.save(path)
    wb.close()
    return path


def _create_large_xlsx_stub(size_mb: float) -> str:
    """Create an XLSX file of approximately `size_mb` megabytes.

    For efficiency, creates a workbook with enough rows to approach
    the target size. Uses string data to inflate the file.
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "LargeSheet"

    # Each row with a formula and a long string ≈ 100-200 bytes compressed
    # We need roughly size_mb * 1024 * 1024 bytes
    target_bytes = int(size_mb * 1024 * 1024)
    row_data = "Engineering calculation data " * 5  # ~150 chars
    rows_needed = min(target_bytes // 200, 10_000)  # Cap for test speed

    ws["A1"] = "Index"
    ws["B1"] = "Data"
    ws["C1"] = "Formula"
    for i in range(2, rows_needed + 2):
        ws[f"A{i}"] = i
        ws[f"B{i}"] = row_data
        ws[f"C{i}"] = f"=A{i}*2"

    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    wb.save(path)
    wb.close()
    return path


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


class TestExtractsFormulasFromSmallFile:
    """Basic formula extraction works for normal-sized files."""

    def test_extracts_formulas_from_small_file(self):
        from scripts.document_intelligence.xlsx_formula_extractor import (
            extract_xlsx_formulas,
        )

        path = _create_small_xlsx(formulas=True)
        try:
            result = extract_xlsx_formulas(path)
            assert result["status"] == "success"
            assert result["formula_count"] >= 3
            # Check formula inventory format
            formulas = result["formulas"]
            assert len(formulas) >= 3
            assert all("cell_ref" in f for f in formulas)
            assert all("sheet" in f for f in formulas)
            assert all("formula" in f for f in formulas)
            # Verify specific formulas
            formula_texts = [f["formula"] for f in formulas]
            assert any("A2+B2" in f for f in formula_texts)
        finally:
            os.unlink(path)

    def test_no_formulas_returns_empty(self):
        from scripts.document_intelligence.xlsx_formula_extractor import (
            extract_xlsx_formulas,
        )

        path = _create_small_xlsx(formulas=False)
        try:
            result = extract_xlsx_formulas(path)
            assert result["status"] == "success"
            assert result["formula_count"] == 0
        finally:
            os.unlink(path)


class TestHandlesLargeFileUpTo50MB:
    """Size limit is raised from 15MB to 50MB."""

    def test_size_limit_is_50mb(self):
        from scripts.document_intelligence.xlsx_formula_extractor import (
            MAX_SIZE_MB,
        )

        assert MAX_SIZE_MB == 50, f"Expected 50MB limit, got {MAX_SIZE_MB}MB"

    def test_accepts_file_under_50mb(self):
        """A file under 50MB should be processed (not skipped)."""
        from scripts.document_intelligence.xlsx_formula_extractor import (
            extract_xlsx_formulas,
        )

        path = _create_small_xlsx(formulas=True)
        try:
            # Small file is well under 50MB
            result = extract_xlsx_formulas(path)
            assert result["status"] == "success"
        finally:
            os.unlink(path)

    def test_rejects_file_over_50mb(self):
        """A file over 50MB should be rejected with clear message."""
        from scripts.document_intelligence.xlsx_formula_extractor import (
            extract_xlsx_formulas,
        )

        path = _create_small_xlsx(formulas=True)
        try:
            # Mock the file size to be >50MB
            with patch("pathlib.Path.stat") as mock_stat:
                mock_stat.return_value = MagicMock(
                    st_size=55 * 1024 * 1024  # 55MB
                )
                result = extract_xlsx_formulas(path)
                assert result["status"] == "skipped"
                assert "50" in result.get("reason", "")
        finally:
            os.unlink(path)


class TestMemoryEfficientStreaming:
    """Verify streaming/chunked approach via openpyxl read_only mode."""

    def test_uses_read_only_mode(self):
        """Extraction must use openpyxl read_only=True for data pass."""
        from scripts.document_intelligence.xlsx_formula_extractor import (
            extract_xlsx_formulas,
        )

        path = _create_small_xlsx(formulas=True)
        try:
            with patch(
                "scripts.document_intelligence.xlsx_formula_extractor.load_workbook"
            ) as mock_load:
                # Set up mock to return a workbook-like object
                mock_wb = MagicMock()
                mock_wb.sheetnames = ["Sheet1"]
                mock_ws = MagicMock()
                mock_ws.iter_rows.return_value = []
                mock_wb.__getitem__ = MagicMock(return_value=mock_ws)
                mock_load.return_value = mock_wb

                extract_xlsx_formulas(path)

                # Verify read_only=True was used in at least one call
                calls = mock_load.call_args_list
                assert len(calls) >= 1
                # First call (data pass) should use read_only=True
                _, kwargs = calls[0]
                assert kwargs.get("read_only") is True, (
                    f"First load_workbook call should use read_only=True, "
                    f"got kwargs: {kwargs}"
                )
        finally:
            os.unlink(path)


class TestErrorOnCorruptXlsx:
    """Graceful error handling for corrupt/invalid XLSX files."""

    def test_error_on_corrupt_xlsx(self):
        from scripts.document_intelligence.xlsx_formula_extractor import (
            extract_xlsx_formulas,
        )

        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.write(fd, b"NOT A VALID XLSX FILE")
        os.close(fd)
        try:
            result = extract_xlsx_formulas(path)
            assert result["status"] == "error"
            assert "reason" in result
        finally:
            os.unlink(path)

    def test_error_on_missing_file(self):
        from scripts.document_intelligence.xlsx_formula_extractor import (
            extract_xlsx_formulas,
        )

        result = extract_xlsx_formulas("/nonexistent/file.xlsx")
        assert result["status"] == "error"
        assert "not found" in result.get("reason", "").lower()
