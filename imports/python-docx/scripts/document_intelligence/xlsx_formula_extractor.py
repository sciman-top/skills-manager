#!/usr/bin/env python3
# /// script
# requires-python = ">=3.10"
# dependencies = [
#     "openpyxl",
# ]
# ///
"""XLSX formula extraction with 50MB size limit and streaming (#1619).

Extracts all formulas from XLSX files using openpyxl in memory-efficient
read_only mode. Raises the previous 15MB limit to 50MB for large
engineering spreadsheets (structural calculations, cost models).

Key changes from the 15MB baseline:
  - Size limit raised from 15MB to 50MB
  - Pass 1 (data values) uses read_only=True for streaming
  - Pass 2 (formulas) uses read_only=False (required by openpyxl to read formulas)
  - JSON output with formula inventory (sheet, cell_ref, formula text)

Usage:
    uv run scripts/document_intelligence/xlsx_formula_extractor.py --input <file.xlsx>

API:
    from scripts.document_intelligence.xlsx_formula_extractor import extract_xlsx_formulas
    result = extract_xlsx_formulas("path/to/large.xlsx")
"""

import argparse
import json
import sys
import time
from pathlib import Path
from typing import Any, Dict, List

# Ensure repo root on PYTHONPATH
_REPO_ROOT = str(Path(__file__).resolve().parents[2])
if _REPO_ROOT not in sys.path:
    sys.path.insert(0, _REPO_ROOT)

from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Size limit — raised from 15MB (#1619)
# ---------------------------------------------------------------------------

MAX_SIZE_MB = 50
_MAX_SIZE_BYTES = MAX_SIZE_MB * 1024 * 1024


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _cell_ref_str(col: int, row: int) -> str:
    """Convert 1-based column and row to Excel ref like 'A1'."""
    result = ""
    c = col
    while c > 0:
        c, remainder = divmod(c - 1, 26)
        result = chr(65 + remainder) + result
    return f"{result}{row}"


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def extract_xlsx_formulas(
    filepath: str,
    max_size_mb: int = MAX_SIZE_MB,
) -> Dict[str, Any]:
    """Extract all formulas from an XLSX file.

    Uses a dual-pass approach:
      Pass 1 (read_only=True):  cached computed values (memory-efficient streaming)
      Pass 2 (read_only=False): formula strings (required by openpyxl)

    Args:
        filepath: Path to the XLSX file.
        max_size_mb: Maximum file size in MB (default: 50).

    Returns:
        Dict with keys:
          - status: "success" | "skipped" | "error"
          - reason: error/skip reason (if applicable)
          - formula_count: number of formulas found
          - formulas: list of {sheet, cell_ref, formula, cached_value}
          - sheets: list of sheet names
          - file_size_mb: file size in MB
          - extraction_time_s: elapsed seconds
    """
    p = Path(filepath)

    # Guard: file must exist
    if not p.exists():
        return {
            "status": "error",
            "reason": f"File not found: {filepath}",
            "formula_count": 0,
            "formulas": [],
        }

    # Guard: check file size
    file_size = p.stat().st_size
    file_size_mb = round(file_size / (1024 * 1024), 2)
    max_bytes = max_size_mb * 1024 * 1024

    if file_size > max_bytes:
        return {
            "status": "skipped",
            "reason": (
                f"File too large ({file_size_mb:.1f} MB > "
                f"{max_size_mb} MB limit)"
            ),
            "formula_count": 0,
            "formulas": [],
            "file_size_mb": file_size_mb,
        }

    t0 = time.time()

    try:
        # Pass 1: cached values (read_only=True for memory efficiency)
        wb_data = load_workbook(
            str(p), read_only=True, data_only=True
        )
        cached_values: Dict[str, Any] = {}
        sheet_names = list(wb_data.sheetnames)

        for sheet_name in sheet_names:
            ws = wb_data[sheet_name]
            if not hasattr(ws, "iter_rows"):
                continue
            for row in ws.iter_rows(values_only=False):
                for cell in row:
                    val = cell.value
                    if val is not None:
                        key = f"{sheet_name}!{_cell_ref_str(cell.column, cell.row)}"
                        cached_values[key] = val
        wb_data.close()

        # Pass 2: formula strings (read_only=False — required by openpyxl)
        wb_formula = load_workbook(
            str(p), read_only=False, data_only=False
        )
        formulas: List[Dict[str, Any]] = []

        for sheet_name in wb_formula.sheetnames:
            ws = wb_formula[sheet_name]
            if not hasattr(ws, "iter_rows"):
                continue
            for row in ws.iter_rows():
                for cell in row:
                    val = cell.value
                    if isinstance(val, str) and val.startswith("="):
                        ref = _cell_ref_str(cell.column, cell.row)
                        full_key = f"{sheet_name}!{ref}"
                        cached = cached_values.get(full_key)
                        formulas.append({
                            "sheet": sheet_name,
                            "cell_ref": ref,
                            "formula": val,
                            "cached_value": cached,
                        })
        wb_formula.close()

        elapsed = round(time.time() - t0, 3)

        return {
            "status": "success",
            "formula_count": len(formulas),
            "formulas": formulas,
            "sheets": sheet_names,
            "file_size_mb": file_size_mb,
            "extraction_time_s": elapsed,
        }

    except Exception as exc:
        elapsed = round(time.time() - t0, 3)
        return {
            "status": "error",
            "reason": str(exc),
            "formula_count": 0,
            "formulas": [],
            "file_size_mb": file_size_mb,
            "extraction_time_s": elapsed,
        }


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> int:
    parser = argparse.ArgumentParser(
        description="XLSX formula extraction with 50MB limit (#1619)"
    )
    parser.add_argument("--input", required=True, help="Path to XLSX file")
    parser.add_argument(
        "--output", help="Output JSON path (default: stdout)"
    )
    parser.add_argument("--verbose", action="store_true")
    args = parser.parse_args()

    result = extract_xlsx_formulas(args.input)

    if args.output:
        Path(args.output).parent.mkdir(parents=True, exist_ok=True)
        with open(args.output, "w") as f:
            json.dump(result, f, indent=2, default=str)
        print(f"Written: {args.output}")
    else:
        print(json.dumps(result, indent=2, default=str))

    if result["status"] == "success":
        print(
            f"\nExtracted {result['formula_count']} formulas from "
            f"{len(result.get('sheets', []))} sheets "
            f"({result.get('file_size_mb', 0)} MB, "
            f"{result.get('extraction_time_s', 0)}s)",
            file=sys.stderr,
        )
        return 0
    else:
        print(f"\n{result['status']}: {result.get('reason', '')}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    sys.exit(main())
